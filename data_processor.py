"""
授業アンケートデータ処理モジュール

CSVデータの読み込み、クレンジング、スコアリング、集計を行う関数群
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import io


# スコアリング変換マッピング（4件法）
SCORE_MAPPING = {
    # 4点（最高評価）
    "とてもそう思う": 4,
    "当てはまる": 4,
    "とても当てはまる": 4,
    "強くそう思う": 4,

    # 3点
    "そう思う": 3,
    "どちらかといえばそう思う": 3,
    "やや当てはまる": 3,
    "どちらかといえば当てはまる": 3,

    # 2点
    "あまりそう思わない": 2,
    "どちらかといえばそう思わない": 2,
    "あまり当てはまらない": 2,
    "どちらかといえば当てはまらない": 2,

    # 1点（最低評価）
    "思わない": 1,
    "そう思わない": 1,
    "当てはまらない": 1,
    "全く当てはまらない": 1,
    "全くそう思わない": 1,
}


# 除外するメタデータカラム名（これらは質問項目として扱わない）
METADATA_COLUMNS = [
    "Id",
    "id",
    "ID",
    "開始時刻",
    "完了時刻",
    "メール",
    "メールアドレス",
    "Email",
    "email",
    "名前",
    "氏名",
    "Name",
    "name",
    "タイムスタンプ",
    "Timestamp",
    "timestamp",
]

# 必須カラム（科目名識別用）
SUBJECT_COLUMN_PATTERNS = [
    "科目名",
    "科目",
    "教科名",
    "授業名",
]

# 出席番号カラムのパターン
STUDENT_ID_PATTERNS = [
    "出席番号",
    "学籍番号",
    "学生番号",
]

# 自由記述カラムのパターン
FREE_TEXT_PATTERNS = [
    "意見・感想",
    "意見",
    "感想",
    "コメント",
    "自由記述",
    "その他",
    "記入してください",
    "ご記入ください",
]


def convert_to_score(value: str) -> Optional[float]:
    """
    4件法のテキスト回答を数値スコアに変換

    Args:
        value: 回答テキスト

    Returns:
        float: 変換後のスコア（1-4）、変換できない場合はNone
    """
    if pd.isna(value):
        return None

    # 文字列に変換して前後の空白を削除
    value_str = str(value).strip()

    # マッピングテーブルから検索
    return SCORE_MAPPING.get(value_str, None)


def detect_subject_column(df: pd.DataFrame) -> Optional[str]:
    """
    科目名カラムを自動検出

    Args:
        df: データフレーム

    Returns:
        str: 科目名カラム名、見つからない場合はNone
    """
    for col in df.columns:
        for pattern in SUBJECT_COLUMN_PATTERNS:
            if pattern in col:
                return col
    return None


def detect_student_id_column(df: pd.DataFrame) -> Optional[str]:
    """
    出席番号カラムを自動検出

    Args:
        df: データフレーム

    Returns:
        str: 出席番号カラム名、見つからない場合はNone
    """
    for col in df.columns:
        for pattern in STUDENT_ID_PATTERNS:
            if pattern in col:
                return col
    return None


def detect_free_text_column(df: pd.DataFrame) -> Optional[str]:
    """
    自由記述カラムを自動検出

    Args:
        df: データフレーム

    Returns:
        str: 自由記述カラム名、見つからない場合はNone
    """
    # より具体的なパターンから順に検索
    # （「意見」だけでなく「ご記入ください」などの具体的なパターンを優先）
    priority_patterns = ["ご記入ください", "記入してください", "意見・感想", "自由記述"]

    # 優先パターンで検索
    for pattern in priority_patterns:
        for col in df.columns:
            if pattern in col:
                return col

    # その他のパターンで検索
    for col in df.columns:
        for pattern in FREE_TEXT_PATTERNS:
            if pattern in col and pattern not in priority_patterns:
                return col

    return None


def identify_question_columns(df: pd.DataFrame) -> List[str]:
    """
    質問項目カラムを識別（メタデータ、科目名、出席番号、自由記述を除外）

    Args:
        df: データフレーム

    Returns:
        List[str]: 質問項目カラムのリスト
    """
    # 除外カラムを特定
    exclude_cols = set()

    # メタデータカラム
    for col in df.columns:
        if col in METADATA_COLUMNS:
            exclude_cols.add(col)

    # 科目名カラム
    subject_col = detect_subject_column(df)
    if subject_col:
        exclude_cols.add(subject_col)

    # 出席番号カラム
    student_id_col = detect_student_id_column(df)
    if student_id_col:
        exclude_cols.add(student_id_col)

    # 自由記述カラム
    free_text_col = detect_free_text_column(df)
    if free_text_col:
        exclude_cols.add(free_text_col)

    # 質問項目カラムを抽出
    question_cols = [col for col in df.columns if col not in exclude_cols]

    return question_cols


def load_and_process_csv(uploaded_file) -> Tuple[pd.DataFrame, Dict]:
    """
    アップロードされたCSVまたはExcelファイルを読み込み、処理する

    Args:
        uploaded_file: StreamlitのUploadedFileオブジェクト

    Returns:
        Tuple[pd.DataFrame, Dict]: 処理済みデータフレームとメタデータ
    """
    # ファイル名から拡張子を取得
    file_name = uploaded_file.name if hasattr(uploaded_file, 'name') else ''
    file_extension = file_name.lower().split('.')[-1] if '.' in file_name else ''

    # ExcelファイルかCSVファイルかを判定して読み込み
    if file_extension in ['xlsx', 'xls']:
        # Excelファイルを読み込み
        df = pd.read_excel(uploaded_file, engine='openpyxl')
    else:
        # CSVを読み込み（エンコーディングを自動判定）
        try:
            df = pd.read_csv(uploaded_file, encoding='utf-8-sig')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(uploaded_file, encoding='shift-jis')
            except UnicodeDecodeError:
                df = pd.read_csv(uploaded_file, encoding='cp932')

    # カラム検出
    subject_col = detect_subject_column(df)
    student_id_col = detect_student_id_column(df)
    free_text_col = detect_free_text_column(df)
    question_cols = identify_question_columns(df)

    # メタデータを作成
    metadata = {
        'subject_column': subject_col,
        'student_id_column': student_id_col,
        'free_text_column': free_text_col,
        'question_columns': question_cols,
        'total_responses': len(df),
    }

    return df, metadata


def calculate_statistics(df: pd.DataFrame, question_cols: List[str]) -> pd.DataFrame:
    """
    質問項目ごとの統計情報を計算

    Args:
        df: データフレーム
        question_cols: 質問項目カラムのリスト

    Returns:
        pd.DataFrame: 統計情報（質問、平均値、各スコアの回答数）
    """
    stats_list = []

    for question in question_cols:
        # スコアに変換
        scores = df[question].apply(convert_to_score)

        # 有効な回答数
        valid_count = scores.notna().sum()

        if valid_count == 0:
            continue

        # 平均値
        mean_score = scores.mean()

        # 各スコアの分布
        score_counts = scores.value_counts().sort_index()
        count_4 = score_counts.get(4.0, 0)
        count_3 = score_counts.get(3.0, 0)
        count_2 = score_counts.get(2.0, 0)
        count_1 = score_counts.get(1.0, 0)

        stats_list.append({
            '質問項目': question,
            '平均値': mean_score,
            '有効回答数': valid_count,
            '4点の回答数': int(count_4),
            '3点の回答数': int(count_3),
            '2点の回答数': int(count_2),
            '1点の回答数': int(count_1),
        })

    return pd.DataFrame(stats_list)


def get_overall_average(df: pd.DataFrame, question_cols: List[str]) -> float:
    """
    全質問の総合平均点を計算

    Args:
        df: データフレーム
        question_cols: 質問項目カラムのリスト

    Returns:
        float: 総合平均点
    """
    all_scores = []

    for question in question_cols:
        scores = df[question].apply(convert_to_score)
        all_scores.extend(scores.dropna().tolist())

    if len(all_scores) == 0:
        return 0.0

    return np.mean(all_scores)


def extract_free_comments(df: pd.DataFrame, free_text_col: str,
                         exclude_empty: bool = True) -> List[str]:
    """
    自由記述を抽出

    Args:
        df: データフレーム
        free_text_col: 自由記述カラム名
        exclude_empty: 空白や「特になし」を除外するか

    Returns:
        List[str]: 自由記述のリスト
    """
    if not free_text_col or free_text_col not in df.columns:
        return []

    comments = df[free_text_col].tolist()

    if exclude_empty:
        # 空白、NaN、「特になし」「特にありません」などを除外
        exclude_patterns = ['特になし', '特にありません', 'なし', '無し', '']
        comments = [
            str(c).strip() for c in comments
            if pd.notna(c) and str(c).strip() not in exclude_patterns
        ]

    return comments


def create_download_data(stats_df: pd.DataFrame, overall_avg: float,
                        subject_name: str = "全体") -> pd.DataFrame:
    """
    ダウンロード用のExcelデータを作成

    Args:
        stats_df: 統計データフレーム
        overall_avg: 総合平均点
        subject_name: 科目名

    Returns:
        pd.DataFrame: ダウンロード用データ
    """
    # データをコピー
    download_df = stats_df.copy()

    # パーセンテージカラムを追加
    download_df['4点の割合(%)'] = (
        download_df['4点の回答数'] / download_df['有効回答数'] * 100
    ).round(1)
    download_df['3点の割合(%)'] = (
        download_df['3点の回答数'] / download_df['有効回答数'] * 100
    ).round(1)
    download_df['2点の割合(%)'] = (
        download_df['2点の回答数'] / download_df['有効回答数'] * 100
    ).round(1)
    download_df['1点の割合(%)'] = (
        download_df['1点の回答数'] / download_df['有効回答数'] * 100
    ).round(1)

    # 平均値を小数点2桁に丸める
    download_df['平均値'] = download_df['平均値'].round(2)

    # カラムの順序を整理
    download_df = download_df[[
        '質問項目',
        '平均値',
        '有効回答数',
        '4点の回答数',
        '4点の割合(%)',
        '3点の回答数',
        '3点の割合(%)',
        '2点の回答数',
        '2点の割合(%)',
        '1点の回答数',
        '1点の割合(%)',
    ]]

    return download_df


def write_to_template(df: pd.DataFrame, question_cols: List[str],
                     subject_mapping: Optional[Dict[str, List[str]]] = None,
                     placeholders: Optional[Dict[str, str]] = None,
                     template_path: str = "テンプレート.xlsx") -> Tuple[io.BytesIO, Dict]:
    """
    テンプレートExcelファイルに各教科のデータを書き込む

    Args:
        df: 全データを含むデータフレーム
        question_cols: 質問項目カラムのリスト
        subject_mapping: ユーザーが選択した教科と科目のマッピング（教科名 -> 科目名リスト）
        placeholders: テンプレートのプレースホルダー（{Y}, {n}, {MM}など）とその値
        template_path: テンプレートファイルのパス

    Returns:
        Tuple[BytesIO, Dict]: (書き込み済みExcelファイル, マッチング情報)
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    import re

    # テンプレートファイルを読み込む
    wb = openpyxl.load_workbook(template_path)
    ws = wb['概要']

    # テンプレートの6行目から質問項目の名前を読み込む（C列からAF列まで）
    template_questions = []
    for col_idx in range(3, 33):  # C列（3）からAF列（32）まで
        cell = ws.cell(row=6, column=col_idx)
        if cell.value:
            template_questions.append((col_idx, str(cell.value)))

    # マッチング情報を格納する辞書
    match_info = {
        'template_question_count': len(template_questions),
        'data_question_count': len(question_cols),
    }

    # プレースホルダーを置換する関数
    def replace_placeholders(text):
        if placeholders and text:
            for key, value in placeholders.items():
                text = text.replace(f'{{{key}}}', str(value))
        return text

    # 1行目と2行目のプレースホルダーを置換
    for row_idx in [1, 2]:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                cell.value = replace_placeholders(cell.value)

    # 7行目から16行目のB列（質問項目列）のプレースホルダーを置換
    for row_idx in range(7, 17):
        cell = ws.cell(row=row_idx, column=2)  # B列
        if cell.value and isinstance(cell.value, str):
            cell.value = replace_placeholders(cell.value)

    # 教科名のマッピング（テンプレートの行番号）
    subject_row_mapping = {
        '全体': 7,
        '国語': 8,
        '数学': 9,
        '地歴公民': 10,
        '理科': 11,
        '外国語': 12,
        '保健体育': 13,
        '芸術': 14,
        '家庭': 15,
        'SS': 16,
    }

    # 科目カラムを検出
    subject_col = detect_subject_column(df)

    # 全体の統計を計算
    overall_stats = calculate_statistics(df, question_cols)

    # 質問項目数の警告
    match_info['excess_questions'] = question_cols[30:] if len(question_cols) > 30 else []

    # 質問項目のマッピングを作成
    # アップロードされたデータの質問項目とテンプレートの質問項目を照合
    question_mapping = {}  # {テンプレート列番号: データの質問項目インデックス}
    match_details = {}  # マッチング詳細情報 {テンプレート列番号: (マッチタイプ, データ質問文)}

    for template_col_idx, template_question in template_questions:
        matched = False

        # まず完全一致を探す
        for data_idx, data_question in enumerate(question_cols):
            if data_question == template_question:
                question_mapping[template_col_idx] = data_idx
                match_details[template_col_idx] = ("完全一致", data_question)
                matched = True
                break

        # 完全一致がない場合、部分一致を試す
        if not matched:
            for data_idx, data_question in enumerate(question_cols):
                # テンプレート質問がデータ質問に含まれているか
                if template_question in data_question:
                    question_mapping[template_col_idx] = data_idx
                    match_details[template_col_idx] = ("部分一致（テンプレート⊂データ）", data_question)
                    matched = True
                    break
                # データ質問がテンプレート質問に含まれているか
                elif data_question in template_question:
                    question_mapping[template_col_idx] = data_idx
                    match_details[template_col_idx] = ("部分一致（データ⊂テンプレート）", data_question)
                    matched = True
                    break

    # マッチングタイプ別の統計
    exact_matches = sum(1 for match_type, _ in match_details.values() if match_type == "完全一致")
    partial_matches = len(match_details) - exact_matches

    # マッチング情報を格納
    match_info['total_matches'] = len(question_mapping)
    match_info['exact_matches'] = exact_matches
    match_info['partial_matches'] = partial_matches

    # マッピングされなかった質問項目
    unmapped_template = [(col_idx, q) for col_idx, q in template_questions if col_idx not in question_mapping]
    match_info['unmapped_questions'] = unmapped_template

    # 部分一致した質問項目の詳細
    partial_match_details = []
    for col_idx, (match_type, data_question) in match_details.items():
        if match_type != "完全一致":
            template_q = next(q for c, q in template_questions if c == col_idx)
            partial_match_details.append({
                'column': col_idx,
                'match_type': match_type,
                'template_question': template_q,
                'data_question': data_question
            })
    match_info['partial_match_details'] = partial_match_details

    # 全体データを書き込み（7行目）
    row_idx = subject_row_mapping['全体']
    avg_values = overall_stats['平均値'].tolist()

    for template_col_idx, data_idx in question_mapping.items():
        if data_idx < len(avg_values):
            ws.cell(row=row_idx, column=template_col_idx, value=round(avg_values[data_idx], 2))

    # AG列（列33）に平均値を書き込み
    overall_avg = np.mean([avg_values[data_idx] for data_idx in question_mapping.values() if data_idx < len(avg_values)])
    ws.cell(row=row_idx, column=33, value=round(overall_avg, 2))

    # AH列（列34）にサンプルサイズ（回答数）を書き込み
    ws.cell(row=row_idx, column=34, value=int(len(df)))

    # 各教科のデータを処理して書き込む
    if subject_col and subject_col in df.columns:
        # ユーザーが選択したマッピングがある場合はそれを使用
        if subject_mapping:
            # ユーザー選択のマッピングを使用
            for template_subject, row_idx in subject_row_mapping.items():
                if template_subject == '全体':
                    continue

                # ユーザーが選択した科目のリストを取得
                matched_subjects = subject_mapping.get(template_subject, [])

                # マッチした教科のデータをフィルタリング
                if matched_subjects:
                    subject_df = df[df[subject_col].isin(matched_subjects)]

                    if len(subject_df) == 0:
                        continue

                    # 統計を計算
                    subject_stats = calculate_statistics(subject_df, question_cols)

                    # データを書き込み（質問項目のマッピングを使用）
                    avg_values = subject_stats['平均値'].tolist()
                    for template_col_idx, data_idx in question_mapping.items():
                        if data_idx < len(avg_values):
                            ws.cell(row=row_idx, column=template_col_idx, value=round(avg_values[data_idx], 2))

                    # AG列（列33）に平均値を書き込み
                    subject_avg = np.mean([avg_values[data_idx] for data_idx in question_mapping.values() if data_idx < len(avg_values)])
                    ws.cell(row=row_idx, column=33, value=round(subject_avg, 2))

                    # AH列（列34）にサンプルサイズ（回答数）を書き込み
                    ws.cell(row=row_idx, column=34, value=int(len(subject_df)))
        else:
            # デフォルトの自動マッピング（後方互換性のため）
            # 教科名の部分一致用キーワード
            subject_keywords = {
                '国語': ['国語', 'こくご'],
                '数学': ['数学', 'すうがく'],
                '地歴公民': ['地理', '歴史', '公民', '地歴', '社会'],
                '理科': ['理科', '物理', '化学', '生物', '地学'],
                '外国語': ['英語', '外国語', 'English'],
                '保健体育': ['保健', '体育', 'たいいく'],
                '芸術': ['音楽', '美術', '書道', '芸術'],
                '家庭': ['家庭', 'かてい'],
                '情報': ['情報', 'じょうほう'],
            }

            # データに含まれる教科名を取得
            unique_subjects = df[subject_col].unique()

            # 各テンプレート教科について処理
            for template_subject, row_idx in subject_row_mapping.items():
                if template_subject == '全体':
                    continue

                # 部分一致で教科を検索
                matched_subjects = []
                keywords = subject_keywords.get(template_subject, [])

                for actual_subject in unique_subjects:
                    if pd.isna(actual_subject):
                        continue

                    actual_subject_str = str(actual_subject)

                    # 完全一致をチェック
                    if actual_subject_str == template_subject:
                        matched_subjects.append(actual_subject)
                        continue

                    # キーワードによる部分一致をチェック
                    for keyword in keywords:
                        if keyword in actual_subject_str:
                            matched_subjects.append(actual_subject)
                            break

                # マッチした教科のデータをフィルタリング
                if matched_subjects:
                    subject_df = df[df[subject_col].isin(matched_subjects)]

                    if len(subject_df) == 0:
                        continue

                    # 統計を計算
                    subject_stats = calculate_statistics(subject_df, question_cols)

                    # データを書き込み（質問項目のマッピングを使用）
                    avg_values = subject_stats['平均値'].tolist()
                    for template_col_idx, data_idx in question_mapping.items():
                        if data_idx < len(avg_values):
                            ws.cell(row=row_idx, column=template_col_idx, value=round(avg_values[data_idx], 2))

                    # AG列（列33）に平均値を書き込み
                    subject_avg = np.mean([avg_values[data_idx] for data_idx in question_mapping.values() if data_idx < len(avg_values)])
                    ws.cell(row=row_idx, column=33, value=round(subject_avg, 2))

                    # AH列（列34）にサンプルサイズ（回答数）を書き込み
                    ws.cell(row=row_idx, column=34, value=int(len(subject_df)))

    # ========================================
    # 各教科の詳細シートを生成
    # ========================================
    created_sheets = []
    if subject_col and subject_col in df.columns and subject_mapping:
        for template_subject, matched_subjects in subject_mapping.items():
            # 教科に科目が割り当てられている場合のみシートを作成
            if not matched_subjects:
                continue

            # 新しいシートを作成
            ws_subject = wb.create_sheet(title=template_subject)
            created_sheets.append({
                'sheet_name': template_subject,
                'subject_count': len(matched_subjects),
                'subjects': matched_subjects
            })

            # 概要シートの1-6行目（ヘッダー部分）をコピー
            for row_idx in range(1, 7):
                for col_idx in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row_idx, col_idx)
                    target_cell = ws_subject.cell(row_idx, col_idx)

                    # 値をコピー
                    target_cell.value = source_cell.value

                    # スタイルをコピー（フォント、塗りつぶし、罫線など）
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()

            # 列幅をコピー
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if col_letter in ws.column_dimensions:
                    ws_subject.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

            # 行の高さをコピー（1-6行目）
            for row_idx in range(1, 7):
                if row_idx in ws.row_dimensions:
                    if ws.row_dimensions[row_idx].height:
                        ws_subject.row_dimensions[row_idx].height = ws.row_dimensions[row_idx].height

            # 結合セルをコピー（1-6行目に含まれるもの）
            for merged_cell_range in ws.merged_cells.ranges:
                # 結合セルの範囲を取得
                min_row = merged_cell_range.min_row
                max_row = merged_cell_range.max_row
                min_col = merged_cell_range.min_col
                max_col = merged_cell_range.max_col

                # 1-6行目に含まれる結合セルのみコピー
                if max_row <= 6:
                    ws_subject.merge_cells(
                        start_row=min_row,
                        start_column=min_col,
                        end_row=max_row,
                        end_column=max_col
                    )

            # プレースホルダーを置換（1-2行目）
            for row_idx in [1, 2]:
                for col_idx in range(1, ws_subject.max_column + 1):
                    cell = ws_subject.cell(row_idx, col_idx)
                    if cell.value and isinstance(cell.value, str):
                        cell.value = replace_placeholders(cell.value)

            # 7行目以降のデータ行: テンプレートの7行目のフォーマットをコピー
            template_data_row = 7

            # 7行目: 教科全体の統計
            current_row = 7

            # テンプレートの7行目からスタイルをコピー
            for col_idx in range(1, ws.max_column + 1):
                source_cell = ws.cell(template_data_row, col_idx)
                target_cell = ws_subject.cell(current_row, col_idx)

                # スタイルをコピー
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()

            # 行の高さもコピー
            if template_data_row in ws.row_dimensions:
                if ws.row_dimensions[template_data_row].height:
                    ws_subject.row_dimensions[current_row].height = ws.row_dimensions[template_data_row].height

            # A列に教科全体の名前を設定
            ws_subject.cell(current_row, 1, value=f"{template_subject}全体")

            # B列に「R{Y} 第{n}回」を設定（プレースホルダー置換済み）
            template_b_value = ws.cell(template_data_row, 2).value
            if template_b_value:
                replaced_value = replace_placeholders(str(template_b_value))
                ws_subject.cell(current_row, 2, value=replaced_value)

            # 教科全体のデータを取得
            subject_df = df[df[subject_col].isin(matched_subjects)]
            if len(subject_df) > 0:
                subject_stats = calculate_statistics(subject_df, question_cols)
                avg_values = subject_stats['平均値'].tolist()

                # 質問項目のマッピングを使用してデータを書き込み
                for template_col_idx, data_idx in question_mapping.items():
                    if data_idx < len(avg_values):
                        ws_subject.cell(current_row, template_col_idx, value=round(avg_values[data_idx], 2))

                # AG列（列33）に平均値を書き込み
                subject_avg = np.mean([avg_values[data_idx] for data_idx in question_mapping.values() if data_idx < len(avg_values)])
                ws_subject.cell(current_row, 33, value=round(subject_avg, 2))

                # AH列（列34）にサンプルサイズ（回答数）を書き込み
                ws_subject.cell(current_row, 34, value=int(len(subject_df)))

            current_row += 1

            # 各科目の統計を書き込み
            for subject_name in matched_subjects:
                # テンプレートの8行目のスタイルをコピー（7行目の次の行）
                template_style_row = 8
                for col_idx in range(1, ws.max_column + 1):
                    source_cell = ws.cell(template_style_row, col_idx)
                    target_cell = ws_subject.cell(current_row, col_idx)

                    # スタイルをコピー
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()

                # 行の高さもコピー
                if template_style_row in ws.row_dimensions:
                    if ws.row_dimensions[template_style_row].height:
                        ws_subject.row_dimensions[current_row].height = ws.row_dimensions[template_style_row].height

                # A列に科目名を設定
                ws_subject.cell(current_row, 1, value=subject_name)

                # B列に「R{Y} 第{n}回」を設定（プレースホルダー置換済み）
                template_b_value = ws.cell(template_style_row, 2).value
                if template_b_value:
                    replaced_value = replace_placeholders(str(template_b_value))
                    ws_subject.cell(current_row, 2, value=replaced_value)

                # 科目ごとのデータを取得
                subject_only_df = df[df[subject_col] == subject_name]
                if len(subject_only_df) > 0:
                    subject_only_stats = calculate_statistics(subject_only_df, question_cols)
                    avg_values = subject_only_stats['平均値'].tolist()

                    # 質問項目のマッピングを使用してデータを書き込み
                    for template_col_idx, data_idx in question_mapping.items():
                        if data_idx < len(avg_values):
                            ws_subject.cell(current_row, template_col_idx, value=round(avg_values[data_idx], 2))

                    # AG列（列33）に平均値を書き込み
                    subject_only_avg = np.mean([avg_values[data_idx] for data_idx in question_mapping.values() if data_idx < len(avg_values)])
                    ws_subject.cell(current_row, 33, value=round(subject_only_avg, 2))

                    # AH列（列34）にサンプルサイズ（回答数）を書き込み
                    ws_subject.cell(current_row, 34, value=int(len(subject_only_df)))

                current_row += 1

    # 作成したシートの情報をmatch_infoに追加
    match_info['created_sheets'] = created_sheets

    # BytesIOに書き込み
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, match_info


def create_integrated_excel(df: pd.DataFrame, question_cols: List[str]) -> io.BytesIO:
    """
    すべてのデータを統合したExcelファイルを生成

    Args:
        df: 全データを含むデータフレーム
        question_cols: 質問項目カラムのリスト

    Returns:
        BytesIO: 生成されたExcelファイル
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 科目カラムを検出
    subject_col = detect_subject_column(df)

    # ========================================
    # 全体シートを作成
    # ========================================
    ws_overall = wb.create_sheet(title="全体")

    # ヘッダー行を作成
    headers = ['質問項目', '平均値', 'サンプルサイズ', '4点', '3点', '2点', '1点']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_overall.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # 全体の統計を計算
    overall_stats = calculate_statistics(df, question_cols)

    # データ行を書き込み
    for row_idx, (_, row) in enumerate(overall_stats.iterrows(), 2):
        ws_overall.cell(row=row_idx, column=1, value=row['質問項目']).border = border
        ws_overall.cell(row=row_idx, column=2, value=round(row['平均値'], 2)).border = border
        ws_overall.cell(row=row_idx, column=3, value=int(row['有効回答数'])).border = border
        ws_overall.cell(row=row_idx, column=4, value=int(row['4点の回答数'])).border = border
        ws_overall.cell(row=row_idx, column=5, value=int(row['3点の回答数'])).border = border
        ws_overall.cell(row=row_idx, column=6, value=int(row['2点の回答数'])).border = border
        ws_overall.cell(row=row_idx, column=7, value=int(row['1点の回答数'])).border = border

        # 平均値のセルを中央揃え
        ws_overall.cell(row=row_idx, column=2).alignment = center_alignment
        ws_overall.cell(row=row_idx, column=3).alignment = center_alignment
        ws_overall.cell(row=row_idx, column=4).alignment = center_alignment
        ws_overall.cell(row=row_idx, column=5).alignment = center_alignment
        ws_overall.cell(row=row_idx, column=6).alignment = center_alignment
        ws_overall.cell(row=row_idx, column=7).alignment = center_alignment

    # 列幅を調整
    ws_overall.column_dimensions['A'].width = 50
    ws_overall.column_dimensions['B'].width = 12
    ws_overall.column_dimensions['C'].width = 15
    ws_overall.column_dimensions['D'].width = 10
    ws_overall.column_dimensions['E'].width = 10
    ws_overall.column_dimensions['F'].width = 10
    ws_overall.column_dimensions['G'].width = 10

    # サマリー情報を追加（最下部に）
    summary_row = len(overall_stats) + 3
    ws_overall.cell(row=summary_row, column=1, value="全体平均").font = Font(bold=True)
    overall_avg = np.mean(overall_stats['平均値'])
    ws_overall.cell(row=summary_row, column=2, value=round(overall_avg, 2)).font = Font(bold=True)
    ws_overall.cell(row=summary_row, column=3, value=int(len(df))).font = Font(bold=True)

    # ========================================
    # 各教科のシートを作成
    # ========================================
    if subject_col and subject_col in df.columns:
        # 教科名のグループ化用キーワード
        subject_groups = {
            '国語': ['国語', 'こくご', '古典'],
            '数学': ['数学', 'すうがく'],
            '地歴公民': ['地理', '歴史', '公民', '地歴', '社会', '倫理', '政経'. '地理', '世界史'],
            '理科': ['理科', '物理', '化学', '生物', '地学'],
            '外国語': ['英語', '外国語', 'English', '英'],
            '保健体育': ['保健', '体育', 'たいいく'],
            '芸術': ['音楽', '美術', '書道', '芸術'],
            '家庭': ['家庭', 'かてい'],
            '情報': ['情報', 'じょうほう', 'Water', 'WS', 'SSP', 'SS'],
        }

        # データに含まれる科目名を取得
        unique_subjects = sorted([str(s) for s in df[subject_col].unique() if pd.notna(s)])

        # 各教科グループについて処理
        for group_name, keywords in subject_groups.items():
            # このグループに属する科目を検索
            matched_subjects = []
            for subject in unique_subjects:
                # 完全一致または部分一致をチェック
                if subject == group_name:
                    matched_subjects.append(subject)
                else:
                    for keyword in keywords:
                        if keyword in subject:
                            matched_subjects.append(subject)
                            break

            # マッチした科目がない場合はスキップ
            if not matched_subjects:
                continue

            # 教科シートを作成
            ws_subject = wb.create_sheet(title=group_name)

            # ヘッダー行を作成
            for col_idx, header in enumerate(headers, 1):
                cell = ws_subject.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            current_row = 2

            # 教科全体の統計
            subject_df = df[df[subject_col].isin(matched_subjects)]
            if len(subject_df) > 0:
                # サブヘッダー（教科全体）
                cell = ws_subject.cell(row=current_row, column=1, value=f"{group_name}全体")
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.border = border

                subject_stats = calculate_statistics(subject_df, question_cols)
                subject_avg = np.mean(subject_stats['平均値'])

                ws_subject.cell(row=current_row, column=2, value=round(subject_avg, 2)).border = border
                ws_subject.cell(row=current_row, column=2).fill = subheader_fill
                ws_subject.cell(row=current_row, column=2).alignment = center_alignment

                ws_subject.cell(row=current_row, column=3, value=int(len(subject_df))).border = border
                ws_subject.cell(row=current_row, column=3).fill = subheader_fill
                ws_subject.cell(row=current_row, column=3).alignment = center_alignment

                # 4点〜1点のセルも塗りつぶし
                for col_idx in range(4, 8):
                    cell = ws_subject.cell(row=current_row, column=col_idx)
                    cell.fill = subheader_fill
                    cell.border = border

                current_row += 1

                # 質問項目ごとの詳細
                for _, row in subject_stats.iterrows():
                    ws_subject.cell(row=current_row, column=1, value=row['質問項目']).border = border
                    ws_subject.cell(row=current_row, column=2, value=round(row['平均値'], 2)).border = border
                    ws_subject.cell(row=current_row, column=3, value=int(row['有効回答数'])).border = border
                    ws_subject.cell(row=current_row, column=4, value=int(row['4点の回答数'])).border = border
                    ws_subject.cell(row=current_row, column=5, value=int(row['3点の回答数'])).border = border
                    ws_subject.cell(row=current_row, column=6, value=int(row['2点の回答数'])).border = border
                    ws_subject.cell(row=current_row, column=7, value=int(row['1点の回答数'])).border = border

                    # 中央揃え
                    for col_idx in range(2, 8):
                        ws_subject.cell(row=current_row, column=col_idx).alignment = center_alignment

                    current_row += 1

                # 空行
                current_row += 1

            # 各科目の統計
            for subject_name in matched_subjects:
                subject_only_df = df[df[subject_col] == subject_name]
                if len(subject_only_df) == 0:
                    continue

                # サブヘッダー（科目名）
                cell = ws_subject.cell(row=current_row, column=1, value=subject_name)
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.border = border

                subject_only_stats = calculate_statistics(subject_only_df, question_cols)
                subject_only_avg = np.mean(subject_only_stats['平均値'])

                ws_subject.cell(row=current_row, column=2, value=round(subject_only_avg, 2)).border = border
                ws_subject.cell(row=current_row, column=2).fill = subheader_fill
                ws_subject.cell(row=current_row, column=2).alignment = center_alignment

                ws_subject.cell(row=current_row, column=3, value=int(len(subject_only_df))).border = border
                ws_subject.cell(row=current_row, column=3).fill = subheader_fill
                ws_subject.cell(row=current_row, column=3).alignment = center_alignment

                # 4点〜1点のセルも塗りつぶし
                for col_idx in range(4, 8):
                    cell = ws_subject.cell(row=current_row, column=col_idx)
                    cell.fill = subheader_fill
                    cell.border = border

                current_row += 1

                # 質問項目ごとの詳細
                for _, row in subject_only_stats.iterrows():
                    ws_subject.cell(row=current_row, column=1, value=row['質問項目']).border = border
                    ws_subject.cell(row=current_row, column=2, value=round(row['平均値'], 2)).border = border
                    ws_subject.cell(row=current_row, column=3, value=int(row['有効回答数'])).border = border
                    ws_subject.cell(row=current_row, column=4, value=int(row['4点の回答数'])).border = border
                    ws_subject.cell(row=current_row, column=5, value=int(row['3点の回答数'])).border = border
                    ws_subject.cell(row=current_row, column=6, value=int(row['2点の回答数'])).border = border
                    ws_subject.cell(row=current_row, column=7, value=int(row['1点の回答数'])).border = border

                    # 中央揃え
                    for col_idx in range(2, 8):
                        ws_subject.cell(row=current_row, column=col_idx).alignment = center_alignment

                    current_row += 1

                # 空行
                current_row += 1

            # 列幅を調整
            ws_subject.column_dimensions['A'].width = 50
            ws_subject.column_dimensions['B'].width = 12
            ws_subject.column_dimensions['C'].width = 15
            ws_subject.column_dimensions['D'].width = 10
            ws_subject.column_dimensions['E'].width = 10
            ws_subject.column_dimensions['F'].width = 10
            ws_subject.column_dimensions['G'].width = 10

    # BytesIOに書き込み
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_integrated_raw_data_excel(df: pd.DataFrame) -> io.BytesIO:
    """
    すべての生データを統合したExcelファイルを生成

    Args:
        df: 全データを含むデータフレーム（生データ）

    Returns:
        BytesIO: 生成されたExcelファイル
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 科目カラムを検出
    subject_col = detect_subject_column(df)

    # ========================================
    # 全体シートを作成（すべての生データ）
    # ========================================
    ws_overall = wb.create_sheet(title="全体")

    # データフレームをシートに書き込み
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_overall.cell(row=r_idx, column=c_idx, value=value)

            # ヘッダー行のスタイル
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            else:
                cell.border = border

    # 列幅を自動調整
    for column in ws_overall.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # 最大50文字
        ws_overall.column_dimensions[column_letter].width = adjusted_width

    # ========================================
    # 各教科のシートを作成（教科ごとの生データ）
    # ========================================
    if subject_col and subject_col in df.columns:
        # 教科名のグループ化用キーワード
        subject_groups = {
            '国語': ['国語', 'こくご'],
            '数学': ['数学', 'すうがく'],
            '地歴公民': ['地理', '歴史', '公民', '地歴', '社会', '倫理', '政経'],
            '理科': ['理科', '物理', '化学', '生物', '地学'],
            '外国語': ['英語', '外国語', 'English', '英'],
            '保健体育': ['保健', '体育', 'たいいく'],
            '芸術': ['音楽', '美術', '書道', '芸術'],
            '家庭': ['家庭', 'かてい'],
            '情報': ['情報', 'じょうほう', 'Water', 'WS', '探究', 'SSP', 'SS'],
        }

        # データに含まれる科目名を取得
        unique_subjects = sorted([str(s) for s in df[subject_col].unique() if pd.notna(s)])

        # 各教科グループについて処理
        for group_name, keywords in subject_groups.items():
            # このグループに属する科目を検索
            matched_subjects = []
            for subject in unique_subjects:
                # 完全一致または部分一致をチェック
                if subject == group_name:
                    matched_subjects.append(subject)
                else:
                    for keyword in keywords:
                        if keyword in subject:
                            matched_subjects.append(subject)
                            break

            # マッチした科目がない場合はスキップ
            if not matched_subjects:
                continue

            # 教科シートを作成
            ws_subject = wb.create_sheet(title=group_name)

            # 教科に該当するデータをフィルタリング
            subject_df = df[df[subject_col].isin(matched_subjects)]

            if len(subject_df) == 0:
                continue

            # データフレームをシートに書き込み
            for r_idx, row in enumerate(dataframe_to_rows(subject_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_subject.cell(row=r_idx, column=c_idx, value=value)

                    # ヘッダー行のスタイル
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.border = border
                    else:
                        cell.border = border

            # 列幅を自動調整
            for column in ws_subject.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # 最大50文字
                ws_subject.column_dimensions[column_letter].width = adjusted_width

    # BytesIOに書き込み
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output
